# -*- coding: utf-8 -*-
import sys
import datetime
import pytz
import openpyxl
from icalendar import Calendar, Event, Alarm
import jieba

def handler(file):

    tz = pytz.timezone('Asia/Shanghai')
    # filename = file.name
    # filename = sys.argv[1]
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    title = ws['A1'].value
    print 'sucessful loaded sheet %s' % title

    names_lst = []
    new_namelst = []
    date_lst = []

    for cell in ws['E']:
        if cell.value is not None:
            names_lst.append(cell.value)
    names_lst = names_lst[1:]

    # cut title to get year month
    for i in jieba.cut(title):
        if i.isdigit():
            date_lst.append(i)
    year = int(date_lst[0])
    month = int(date_lst[1])

    # just want the first person in the cell value
    for name in names_lst:
        new_namelst.append(''.join(list(jieba.cut(name))[:1]))
    names_lst = new_namelst
    name_sets = set(names_lst)


    values = []
    for day, subject in enumerate(names_lst):
        day += 1
        dtstart = datetime.datetime(year, month, day, 17, 00, tzinfo=tz)
        dtend = datetime.timedelta(seconds=3600) + dtstart

        event = {
            'summary': subject,
            'dtstart': dtstart,
            'dtend': dtend,
        }
        values.append(event)


    # while(name_sets):
        # name = name_sets.pop()
    for name in name_sets:
        cal = Calendar()
        for row in values:
            if row['summary'] == name:
                event = Event()
                alarm = Alarm()
                event.add('summary', row['summary'])
                event.add('dtstart', row['dtstart'])
                event.add('dtend', row['dtend'])
                alarm.add("action", 'DISPLAY')
                alarm.add("description", 'Reminder')
                alarm.add("trigger", row['dtstart'] - datetime.timedelta(seconds=3600))
                # alarm in mac ios icalenar kind of borings
                # event.add_component(alarm)

                cal.add_component(event)
        if cal.subcomponents:
            name_ics = "%d %d %s.ics" % (year,month,name)
            f = open(name_ics,'wb')
            f.write(cal.to_ical())
            print 'success make %s\'s calendar' % name_ics
            f.close()
        else:
            print 'somethings wrong with %s\'s  calendar' % name

if __name__ == '__main__':
    handlefile = sys.argv[1]
    handler(handlefile)
